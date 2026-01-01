"""
Export Module

Exports topic modeling results to various formats (JSON, CSV, Excel).
Supports exporting topics, documents, sentiment, coherence, and full job data.

Usage:
    exporter = JobExporter(job_data)
    exporter.export_json('output.json')
    exporter.export_csv('output.csv')
    exporter.export_excel('output.xlsx')
"""

import logging
import json
import csv
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Try to import excel libraries
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Install with: pip install openpyxl")


class JobExporter:
    """
    Exports topic modeling job results to various formats.
    """

    def __init__(self, job_data: Dict[str, Any]):
        """
        Initialize exporter with job data.

        Args:
            job_data: Complete job data including topics, parameters, sentiment, etc.
        """
        self.job_data = job_data
        self.job_id = job_data.get('job_id', 'unknown')

    def export_json(self, output_path: str, indent: int = 2) -> str:
        """
        Export complete job data to JSON file.

        Args:
            output_path: Path to output JSON file
            indent: JSON indentation (default: 2)

        Returns:
            Path to created file
        """
        logger.info(f"Exporting job {self.job_id} to JSON: {output_path}")

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.job_data, f, indent=indent, ensure_ascii=False)

            logger.info(f"JSON export completed: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise

    def export_topics_csv(self, output_path: str, top_n_words: int = 20) -> str:
        """
        Export topics to CSV file.

        Args:
            output_path: Path to output CSV file
            top_n_words: Number of top words per topic

        Returns:
            Path to created file
        """
        logger.info(f"Exporting topics to CSV: {output_path}")

        try:
            topics = self.job_data.get('topics', [])

            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Header
                writer.writerow([
                    'Topic Number',
                    'Top Words',
                    'Word Weights',
                    'Num Documents',
                    'Avg Sentiment',
                    'Coherence Score'
                ])

                # Data rows
                sentiments = {s['topic_number']: s for s in self.job_data.get('sentiment_analysis', [])}
                coherence = self.job_data.get('coherence_scores', {})
                per_topic_coherence = coherence.get('per_topic_coherence', [])

                for topic in topics:
                    topic_num = topic['topic_number']
                    words = topic['words'][:top_n_words]
                    weights = topic['weights'][:top_n_words]

                    # Format words and weights
                    words_str = ', '.join(words)
                    weights_str = ', '.join([f"{w:.4f}" for w in weights])

                    # Get sentiment if available
                    topic_sentiment = sentiments.get(topic_num, {})
                    avg_sentiment = topic_sentiment.get('avg_sentiment', 'N/A')
                    if isinstance(avg_sentiment, float):
                        avg_sentiment = f"{avg_sentiment:.4f}"

                    # Get coherence if available
                    topic_coherence = 'N/A'
                    if topic_num < len(per_topic_coherence):
                        topic_coherence = f"{per_topic_coherence[topic_num]:.4f}"

                    writer.writerow([
                        topic_num,
                        words_str,
                        weights_str,
                        topic.get('num_documents', 'N/A'),
                        avg_sentiment,
                        topic_coherence
                    ])

            logger.info(f"Topics CSV export completed: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Topics CSV export failed: {e}")
            raise

    def export_documents_csv(
        self,
        output_path: str,
        include_text: bool = False,
        max_docs: Optional[int] = None
    ) -> str:
        """
        Export document-topic assignments to CSV.

        Args:
            output_path: Path to output CSV file
            include_text: Whether to include document text
            max_docs: Maximum number of documents to export (None = all)

        Returns:
            Path to created file
        """
        logger.info(f"Exporting documents to CSV: {output_path}")

        try:
            document_topics = self.job_data.get('document_topics', [])

            if max_docs:
                document_topics = document_topics[:max_docs]

            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Header
                header = ['Document Index', 'Dominant Topic', 'Dominant Topic Probability']

                # Add topic probability columns
                if document_topics:
                    num_topics = len(document_topics[0].get('topic_probabilities', []))
                    header.extend([f'Topic {i} Prob' for i in range(num_topics)])

                if include_text:
                    header.append('Document Text')

                writer.writerow(header)

                # Data rows
                for doc in document_topics:
                    row = [
                        doc['document_index'],
                        doc['dominant_topic'],
                        f"{doc['dominant_probability']:.4f}"
                    ]

                    # Add topic probabilities
                    probs = doc.get('topic_probabilities', [])
                    row.extend([f"{p:.4f}" for p in probs])

                    if include_text:
                        row.append(doc.get('text', ''))

                    writer.writerow(row)

            logger.info(f"Documents CSV export completed: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Documents CSV export failed: {e}")
            raise

    def export_excel(
        self,
        output_path: str,
        include_documents: bool = True,
        max_docs: Optional[int] = 1000
    ) -> str:
        """
        Export complete job data to Excel file with multiple sheets.

        Args:
            output_path: Path to output Excel file
            include_documents: Whether to include document sheet
            max_docs: Maximum documents to include (default: 1000)

        Returns:
            Path to created file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        logger.info(f"Exporting job {self.job_id} to Excel: {output_path}")

        try:
            from openpyxl import Workbook

            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 1. Summary sheet
            self._create_summary_sheet(wb)

            # 2. Topics sheet
            self._create_topics_sheet(wb)

            # 3. Sentiment sheet (if available)
            if self.job_data.get('sentiment_analysis'):
                self._create_sentiment_sheet(wb)

            # 4. Documents sheet (if requested)
            if include_documents:
                self._create_documents_sheet(wb, max_docs)

            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel export completed: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise

    def _create_summary_sheet(self, wb):
        """Create summary sheet with job metadata."""
        ws = wb.create_sheet("Summary")

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Title
        ws['A1'] = "Topic Modeling Job Summary"
        ws['A1'].font = Font(bold=True, size=14)

        # Metadata
        params = self.job_data.get('parameters', {})
        coherence = self.job_data.get('coherence_scores', {})

        data = [
            ['Job ID', self.job_id],
            ['Channel', self.job_data.get('channel_name', 'N/A')],
            ['Algorithm', params.get('algorithm', 'N/A')],
            ['Number of Topics', params.get('num_topics', 'N/A')],
            ['Number of Documents', self.job_data.get('num_documents', 'N/A')],
            ['Created At', self.job_data.get('created_at', 'N/A')],
            [''],
            ['Coherence Score', coherence.get('coherence_score', 'N/A')],
            ['Coherence Type', coherence.get('coherence_type', 'N/A')],
        ]

        for row_idx, (key, value) in enumerate(data, start=3):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_topics_sheet(self, wb):
        """Create topics sheet with top words and statistics."""
        ws = wb.create_sheet("Topics")

        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['Topic #', 'Top 10 Words', 'Num Docs', 'Avg Sentiment', 'Coherence']

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        topics = self.job_data.get('topics', [])
        sentiments = {s['topic_number']: s for s in self.job_data.get('sentiment_analysis', [])}
        coherence = self.job_data.get('coherence_scores', {})
        per_topic_coherence = coherence.get('per_topic_coherence', [])

        for row_idx, topic in enumerate(topics, start=2):
            topic_num = topic['topic_number']

            ws.cell(row=row_idx, column=1, value=topic_num)
            ws.cell(row=row_idx, column=2, value=', '.join(topic['words'][:10]))
            ws.cell(row=row_idx, column=3, value=topic.get('num_documents', 0))

            # Sentiment
            topic_sentiment = sentiments.get(topic_num, {})
            avg_sentiment = topic_sentiment.get('avg_sentiment', None)
            if avg_sentiment is not None:
                ws.cell(row=row_idx, column=4, value=f"{avg_sentiment:.4f}")

            # Coherence
            if topic_num < len(per_topic_coherence):
                ws.cell(row=row_idx, column=5, value=f"{per_topic_coherence[topic_num]:.4f}")

        # Auto-size columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    def _create_sentiment_sheet(self, wb):
        """Create sentiment analysis sheet."""
        ws = wb.create_sheet("Sentiment Analysis")

        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['Topic #', 'Avg Sentiment', 'Sentiment Std', 'Positive', 'Neutral', 'Negative']

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        sentiments = self.job_data.get('sentiment_analysis', [])

        for row_idx, sentiment in enumerate(sentiments, start=2):
            ws.cell(row=row_idx, column=1, value=sentiment['topic_number'])
            ws.cell(row=row_idx, column=2, value=f"{sentiment['avg_sentiment']:.4f}")
            ws.cell(row=row_idx, column=3, value=f"{sentiment['sentiment_std']:.4f}")
            ws.cell(row=row_idx, column=4, value=sentiment['positive_count'])
            ws.cell(row=row_idx, column=5, value=sentiment['neutral_count'])
            ws.cell(row=row_idx, column=6, value=sentiment['negative_count'])

        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

    def _create_documents_sheet(self, wb, max_docs: Optional[int]):
        """Create documents sheet with topic assignments."""
        ws = wb.create_sheet("Documents")

        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['Doc #', 'Dominant Topic', 'Probability', 'Text Preview']

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        document_topics = self.job_data.get('document_topics', [])

        if max_docs:
            document_topics = document_topics[:max_docs]

        for row_idx, doc in enumerate(document_topics, start=2):
            ws.cell(row=row_idx, column=1, value=doc['document_index'])
            ws.cell(row=row_idx, column=2, value=doc['dominant_topic'])
            ws.cell(row=row_idx, column=3, value=f"{doc['dominant_probability']:.4f}")

            # Text preview (first 100 chars)
            text = doc.get('text', '')
            preview = text[:100] + '...' if len(text) > 100 else text
            ws.cell(row=row_idx, column=4, value=preview)

        # Auto-size columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 80


def export_job_to_json(job_data: Dict[str, Any], output_path: str) -> str:
    """
    Convenience function to export job data to JSON.

    Args:
        job_data: Complete job data
        output_path: Path to output JSON file

    Returns:
        Path to created file
    """
    exporter = JobExporter(job_data)
    return exporter.export_json(output_path)


def export_job_to_csv(
    job_data: Dict[str, Any],
    output_dir: str,
    include_documents: bool = True
) -> Dict[str, str]:
    """
    Convenience function to export job data to multiple CSV files.

    Args:
        job_data: Complete job data
        output_dir: Directory for output files
        include_documents: Whether to export documents

    Returns:
        Dict mapping file type to file path
    """
    exporter = JobExporter(job_data)
    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)

    job_id = job_data.get('job_id', 'unknown')
    files = {}

    # Export topics
    topics_path = output_dir_path / f"{job_id}_topics.csv"
    files['topics'] = exporter.export_topics_csv(str(topics_path))

    # Export documents
    if include_documents:
        docs_path = output_dir_path / f"{job_id}_documents.csv"
        files['documents'] = exporter.export_documents_csv(str(docs_path))

    return files


def export_job_to_excel(
    job_data: Dict[str, Any],
    output_path: str,
    include_documents: bool = True,
    max_docs: int = 1000
) -> str:
    """
    Convenience function to export job data to Excel.

    Args:
        job_data: Complete job data
        output_path: Path to output Excel file
        include_documents: Whether to include documents sheet
        max_docs: Maximum documents to include

    Returns:
        Path to created file
    """
    exporter = JobExporter(job_data)
    return exporter.export_excel(output_path, include_documents, max_docs)
